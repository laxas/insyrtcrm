"""Phase 3 Web UI tests — views, permissions, exports, transitions."""

from __future__ import annotations

import io

import pytest
from django.test import Client
from django.urls import reverse
from openpyxl import load_workbook

from apps.accounts.models import UserPreferences
from apps.activities.models import Activity
from apps.leads.models import StageTransition

from .factories import (
    ActivityFactory,
    CompanyFactory,
    ContactFactory,
    PRBriefingFactory,
    StageFactory,
    UserFactory,
)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def auth_client(user):
    c = Client()
    c.force_login(user)
    return c


# ---------------------------------------------------------------------------
# Auth / permission guard
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_lead_list_requires_login():
    c = Client()
    resp = c.get(reverse("lead-list"))
    assert resp.status_code == 302
    assert "login" in resp["Location"]


@pytest.mark.django_db
def test_lead_list_authenticated():
    user = UserFactory()
    c = auth_client(user)
    resp = c.get(reverse("lead-list"))
    assert resp.status_code == 200


@pytest.mark.django_db
def test_lead_detail_requires_login():
    company = CompanyFactory()
    c = Client()
    resp = c.get(reverse("lead-detail", kwargs={"pk": company.pk}))
    assert resp.status_code == 302


# ---------------------------------------------------------------------------
# Lead list — filter / search / sort
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_lead_list_shows_companies():
    user = UserFactory()
    CompanyFactory(name="AcmeCorp")
    c = auth_client(user)
    resp = c.get(reverse("lead-list"))
    assert resp.status_code == 200
    assert b"AcmeCorp" in resp.content


@pytest.mark.django_db
def test_lead_list_filter_by_name():
    user = UserFactory()
    CompanyFactory(name="TargetCo")
    CompanyFactory(name="OtherCo")
    c = auth_client(user)
    resp = c.get(reverse("lead-list"), {"q": "TargetCo"})
    assert b"TargetCo" in resp.content
    assert b"OtherCo" not in resp.content


@pytest.mark.django_db
def test_lead_list_filter_by_stage():
    user = UserFactory()
    stage = StageFactory(name_de="Kaltakquise")
    CompanyFactory(name="InStage", current_stage=stage)
    CompanyFactory(name="NoStage")
    c = auth_client(user)
    resp = c.get(reverse("lead-list"), {"stage": stage.pk})
    assert b"InStage" in resp.content
    assert b"NoStage" not in resp.content


@pytest.mark.django_db
def test_lead_list_filter_by_owner():
    user = UserFactory()
    owner = UserFactory(username="alice")
    CompanyFactory(name="AliceCo", owner=owner)
    CompanyFactory(name="NobodyCo")
    c = auth_client(user)
    resp = c.get(reverse("lead-list"), {"owner": owner.pk})
    assert b"AliceCo" in resp.content
    assert b"NobodyCo" not in resp.content


@pytest.mark.django_db
def test_lead_list_paginate():
    user = UserFactory()
    for i in range(55):
        CompanyFactory(name=f"Co{i:03d}")
    c = auth_client(user)
    resp = c.get(reverse("lead-list"))
    assert resp.status_code == 200
    # Should be paginated — two pages for 55 companies at page_size=50
    assert b"page=2" in resp.content


# ---------------------------------------------------------------------------
# Lead detail + tabs
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_lead_detail_shows_company():
    user = UserFactory()
    company = CompanyFactory(name="DetailCorp", domain="detail.corp")
    c = auth_client(user)
    resp = c.get(reverse("lead-detail", kwargs={"pk": company.pk}))
    assert resp.status_code == 200
    assert b"DetailCorp" in resp.content


@pytest.mark.django_db
def test_tab_briefing_partial():
    user = UserFactory()
    company = CompanyFactory()
    PRBriefingFactory(company=company, reality_check="Real check text")
    c = auth_client(user)
    resp = c.get(reverse("tab-briefing", kwargs={"pk": company.pk}))
    assert resp.status_code == 200
    assert b"Real check text" in resp.content


@pytest.mark.django_db
def test_tab_contacts_partial():
    user = UserFactory()
    company = CompanyFactory()
    ContactFactory(company=company, full_name="Jane Doe")
    c = auth_client(user)
    resp = c.get(reverse("tab-contacts", kwargs={"pk": company.pk}))
    assert resp.status_code == 200
    assert b"Jane Doe" in resp.content


@pytest.mark.django_db
def test_tab_activities_partial():
    user = UserFactory()
    company = CompanyFactory()
    ActivityFactory(company=company, note="Call note here")
    c = auth_client(user)
    resp = c.get(reverse("tab-activities", kwargs={"pk": company.pk}))
    assert resp.status_code == 200
    assert b"Call note here" in resp.content


@pytest.mark.django_db
def test_tab_history_partial():
    user = UserFactory()
    company = CompanyFactory()
    stage_a = StageFactory(name_de="Anfang")
    stage_b = StageFactory(name_de="Mitte")
    company.transition_to(stage_a, user)
    company.transition_to(stage_b, user, comment="moved on")
    c = auth_client(user)
    resp = c.get(reverse("tab-history", kwargs={"pk": company.pk}))
    assert resp.status_code == 200
    assert b"Mitte" in resp.content
    assert b"moved on" in resp.content


# ---------------------------------------------------------------------------
# Stage transition
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_transition_view_get():
    user = UserFactory()
    company = CompanyFactory()
    StageFactory()
    c = auth_client(user)
    resp = c.get(reverse("lead-transition", kwargs={"pk": company.pk}))
    assert resp.status_code == 200


@pytest.mark.django_db
def test_transition_creates_record():
    user = UserFactory()
    company = CompanyFactory()
    stage = StageFactory(name_de="Angebotsphase")
    c = auth_client(user)
    resp = c.post(
        reverse("lead-transition", kwargs={"pk": company.pk}),
        {"stage": stage.pk, "comment": "Moving up", "reduce_data": ""},
    )
    assert resp.status_code == 302
    company.refresh_from_db()
    assert company.current_stage == stage
    assert StageTransition.objects.filter(company=company, to_stage=stage).exists()


@pytest.mark.django_db
def test_transition_archive_reduces_data():
    user = UserFactory()
    company = CompanyFactory()
    ContactFactory(company=company, full_name="To Be Deleted")
    PRBriefingFactory(company=company, reality_check="sensitive info")
    archive_stage = StageFactory(name_de="Archiv", is_archive=True)
    c = auth_client(user)
    c.post(
        reverse("lead-transition", kwargs={"pk": company.pk}),
        {"stage": archive_stage.pk, "comment": "archiving", "reduce_data": "on"},
    )
    company.refresh_from_db()
    assert company.contacts.count() == 0
    assert company.prbriefing.reality_check == ""


# ---------------------------------------------------------------------------
# Activity logging
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_log_phone_activity():
    user = UserFactory()
    company = CompanyFactory()
    c = auth_client(user)
    resp = c.post(
        reverse("log-activity", kwargs={"company_pk": company.pk, "channel": "PHONE"}),
        {
            "occurred_at": "2026-05-01T10:00",
            "direction": Activity.Direction.OUT,
            "outcome": Activity.Outcome.INTERESTED,
            "note": "Good call",
            "duration_minutes": "5",
        },
    )
    assert resp.status_code == 302
    act = Activity.objects.get(company=company)
    assert act.channel == Activity.Channel.PHONE
    assert act.note == "Good call"
    assert act.duration_seconds == 300


@pytest.mark.django_db
def test_log_email_activity():
    user = UserFactory()
    company = CompanyFactory()
    c = auth_client(user)
    resp = c.post(
        reverse("log-activity", kwargs={"company_pk": company.pk, "channel": "EMAIL"}),
        {
            "occurred_at": "2026-05-01T11:00",
            "direction": Activity.Direction.OUT,
            "subject": "Intro mail",
            "note": "",
        },
    )
    assert resp.status_code == 302
    act = Activity.objects.get(company=company)
    assert act.channel == Activity.Channel.EMAIL
    assert act.outcome == Activity.Outcome.SENT
    assert act.subject == "Intro mail"


@pytest.mark.django_db
def test_log_linkedin_activity():
    user = UserFactory()
    company = CompanyFactory()
    c = auth_client(user)
    resp = c.post(
        reverse("log-activity", kwargs={"company_pk": company.pk, "channel": "LINKEDIN"}),
        {
            "occurred_at": "2026-05-01T12:00",
            "outcome": Activity.Outcome.CONNECTED,
            "note": "Sent connection",
        },
    )
    assert resp.status_code == 302
    act = Activity.objects.get(company=company)
    assert act.channel == Activity.Channel.LINKEDIN
    assert act.direction == Activity.Direction.OUT


@pytest.mark.django_db
def test_log_letter_activity():
    user = UserFactory()
    company = CompanyFactory()
    c = auth_client(user)
    resp = c.post(
        reverse("log-activity", kwargs={"company_pk": company.pk, "channel": "LETTER"}),
        {"occurred_at": "2026-05-01T09:00", "note": ""},
    )
    assert resp.status_code == 302
    act = Activity.objects.get(company=company)
    assert act.channel == Activity.Channel.LETTER
    assert act.outcome == Activity.Outcome.SENT
    assert act.direction == Activity.Direction.OUT


# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_csv_export():
    user = UserFactory()
    CompanyFactory(name="ExportMe", domain="export.me")
    c = auth_client(user)
    resp = c.get(reverse("csv-export"))
    assert resp.status_code == 200
    assert resp["Content-Type"].startswith("text/csv")
    content = resp.content.decode("utf-8")
    assert "ExportMe" in content


# ---------------------------------------------------------------------------
# Letter export (FR-LT-01)
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_letter_export_xlsx():
    user = UserFactory()
    company = CompanyFactory(
        name="LetterCo",
        street="Teststr. 1",
        postcode="80333",
        city="München",
        country="Deutschland",
    )
    ContactFactory(company=company, first_name="Max", last_name="Muster", salutation="Herr")
    c = auth_client(user)
    resp = c.post(reverse("letter-export"), {"company_ids": [str(company.pk)]})
    assert resp.status_code == 200
    assert "spreadsheetml" in resp["Content-Type"]

    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0][0] == "Firma"
    data_row = rows[1]
    assert data_row[0] == "LetterCo"
    assert data_row[5] == "Teststr. 1"


@pytest.mark.django_db
def test_letter_export_stores_ids_in_session():
    user = UserFactory()
    company = CompanyFactory()
    c = auth_client(user)
    c.post(reverse("letter-export"), {"company_ids": [str(company.pk)]})
    session = c.session
    assert str(company.pk) in session.get("letter_export_ids", [])


# ---------------------------------------------------------------------------
# Letter log (FR-LT-02)
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_letter_log_creates_activities():
    user = UserFactory()
    c1 = CompanyFactory()
    c2 = CompanyFactory()
    c = auth_client(user)
    # First export to populate session
    c.post(reverse("letter-export"), {"company_ids": [str(c1.pk), str(c2.pk)]})
    resp = c.post(reverse("letter-log"))
    assert resp.status_code == 302
    assert Activity.objects.filter(channel=Activity.Channel.LETTER).count() == 2


@pytest.mark.django_db
def test_letter_log_without_session_redirects():
    user = UserFactory()
    c = auth_client(user)
    resp = c.post(reverse("letter-log"))
    assert resp.status_code == 302
    assert Activity.objects.filter(channel=Activity.Channel.LETTER).count() == 0


# ---------------------------------------------------------------------------
# Column preferences
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_save_columns_updates_prefs():
    user = UserFactory()
    c = auth_client(user)
    resp = c.post(
        reverse("save-columns"),
        {"columns": ["name", "stage", "priority"]},
    )
    assert resp.status_code == 302
    prefs = UserPreferences.objects.get(user=user)
    assert prefs.list_columns == ["name", "stage", "priority"]


@pytest.mark.django_db
def test_save_columns_ignores_invalid_keys():
    user = UserFactory()
    c = auth_client(user)
    c.post(
        reverse("save-columns"),
        {"columns": ["name", "INVALID_KEY"]},
    )
    prefs = UserPreferences.objects.get(user=user)
    assert "INVALID_KEY" not in prefs.list_columns
    assert "name" in prefs.list_columns


@pytest.mark.django_db
def test_get_list_columns_falls_back_to_defaults():
    user = UserFactory()
    prefs = UserPreferences.objects.create(user=user, list_columns=[])
    from apps.accounts.models import DEFAULT_LIST_COLUMNS

    assert prefs.get_list_columns() == DEFAULT_LIST_COLUMNS
